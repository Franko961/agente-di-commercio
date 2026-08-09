import base64
import io
from calendar import monthrange
from datetime import date, datetime, timezone

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.italian_holidays import is_italian_holiday
from core.utils import local_date_str, LOCAL_TZ
from services.export_service import sanitize_cell_text
from services.leave_request_service import LEAVE_TYPE_LABELS


def _local_time_str(iso_ts: str) -> str:
    """Come local_date_str (core.utils), ma restituisce solo l'orario
    'HH:MM' in ora italiana — usato dal foglio Dettaglio, dove l'orologio a
    muro (non l'istante UTC salvato) è quello che interessa a chi elabora
    le buste paga."""
    dt = datetime.fromisoformat(iso_ts.replace("Z", "+00:00"))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(LOCAL_TZ).strftime("%H:%M")

# Stessi colori (senza #) e stesse etichette già usati nella griglia live
# (frontend/src/pages/Presenze.jsx ALL_TYPE_COLORS/ALL_TYPE_LABELS):
# duplicati qui deliberatamente, stesso principio già adottato per
# core/italian_holidays.py (JS e Python restano runtime separati, niente
# modulo condiviso tra frontend e backend).
TYPE_LABELS = {
    "ferie": "Ferie", "permesso": "Permesso", "malattia": "Malattia",
    "smartworking": "Smartworking", "trasferta": "Trasferta",
    "straordinari": "Straordinari", "reperibilita": "Reperibilità",
}
TYPE_COLORS = {
    "ferie": "FF5A00", "permesso": "0A192F", "malattia": "DC2626",
    "smartworking": "D97706", "trasferta": "78350F",
    "straordinari": "DB2777", "reperibilita": "6366F1",
}
PRESENTE_COLOR = "16A34A"

_MONTH_LABELS_IT = (
    "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
    "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre",
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color="FF" + hex_color, end_color="FF" + hex_color, fill_type="solid")


def _is_ferie_excluded_day(mode: str, day: date) -> bool:
    """Stessa logica di isFerieExcludedDay in Presenze.jsx e di
    _days_in_year in leave_request_service.py: un giorno di FERIE non va
    considerato "consumato" in base a ferie_count_mode dell'account."""
    if mode == "calendario":
        return False
    if mode == "lavorativi":
        return day.weekday() >= 5  # sabato=5, domenica=6
    return day.weekday() == 6 or is_italian_holiday(day)  # "festivita"


def _add_logo(ws, company_logo: str) -> None:
    """Decodifica il data URL del logo aziendale e lo ancora in alto a
    sinistra, ridimensionato a un'altezza fissa mantenendo le proporzioni.
    Se il logo non è impostato o è malformato, il file si genera comunque
    senza immagine — un logo corrotto non deve far fallire l'intero export
    del cartellino."""
    if not company_logo:
        return
    try:
        _, b64data = company_logo.split(",", 1)
        logo_bytes = base64.b64decode(b64data)
        xl_img = XLImage(io.BytesIO(logo_bytes))
        target_height = 50
        scale = target_height / xl_img.height
        xl_img.height = target_height
        xl_img.width = round(xl_img.width * scale)
        ws.add_image(xl_img, "A1")
    except Exception:
        pass


def build_attendance_workbook(
    month: str,
    company_name: str,
    company_logo: str,
    ferie_count_mode: str,
    leave_requests: list,
    sessions: list,
) -> Workbook:
    """Cartellino del mese richiesto (AAAA-MM) come due fogli: "Cartellino"
    (vista a griglia dipendenti×giorni colorata come Presenze.jsx, con logo
    e totali di riga) e "Dettaglio" (stessa lista piatta per sessione/
    assenza già prodotta dal precedente export CSV, per chi ha bisogno
    degli orari esatti di entrata/uscita). `leave_requests` sono le
    richieste APPROVATE che si sovrappongono al mese (qualunque tipo,
    inclusi quelli in ADMIN_LEAVE_TYPES); `sessions` sono le sessioni
    presenze CHIUSE il cui giorno solare italiano ricade nel mese —
    entrambe già filtrate dal chiamante (vedi attendance_service.export_xlsx)."""
    year, mon = (int(p) for p in month.split("-"))
    days_in_month = monthrange(year, mon)[1]
    month_label = f"{_MONTH_LABELS_IT[mon - 1]} {year}"

    requests_by_employee: dict = {}
    for r in leave_requests:
        name = r.get("employee_name") or ""
        requests_by_employee.setdefault(name, []).append(r)

    hours_by_key: dict = {}
    for s in sessions:
        name = s.get("employee_name") or ""
        day_iso = local_date_str(s["clock_in"])
        h = (datetime.fromisoformat(s["clock_out"]) - datetime.fromisoformat(s["clock_in"])).total_seconds() / 3600
        hours_by_key[(name, day_iso)] = hours_by_key.get((name, day_iso), 0) + h

    names = sorted(({n for n in requests_by_employee} | {n for n, _ in hours_by_key}) - {""})

    wb = Workbook()
    ws = wb.active
    ws.title = "Cartellino"

    _add_logo(ws, company_logo)

    ws.cell(row=1, column=4, value=sanitize_cell_text(f"Cartellino presenze — {company_name}")).font = Font(bold=True, size=14)
    ws.cell(row=2, column=4, value=month_label).font = Font(size=11, italic=True)

    header_row = 5
    ws.cell(row=header_row, column=1, value="Dipendente").font = Font(bold=True)
    for day in range(1, days_in_month + 1):
        ws.cell(row=header_row, column=1 + day, value=day).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(1 + day)].width = 4
    totals_col = 1 + days_in_month + 1
    ferie_col = totals_col + 1
    ws.cell(row=header_row, column=totals_col, value="Ore totali").font = Font(bold=True)
    ws.cell(row=header_row, column=ferie_col, value="Giorni ferie").font = Font(bold=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions[get_column_letter(totals_col)].width = 12
    ws.column_dimensions[get_column_letter(ferie_col)].width = 12
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    row = header_row + 1
    for name in names:
        ws.cell(row=row, column=1, value=sanitize_cell_text(name))
        total_hours = 0.0
        total_ferie_days = 0
        for day in range(1, days_in_month + 1):
            iso = f"{month}-{day:02d}"
            d = date(year, mon, day)
            match = None
            for r in requests_by_employee.get(name, []):
                if r["date_from"] <= iso <= r["date_to"]:
                    if r["type"] == "ferie" and _is_ferie_excluded_day(ferie_count_mode, d):
                        continue
                    match = r
                    break
            hours = hours_by_key.get((name, iso))
            cell = ws.cell(row=row, column=1 + day)
            if match:
                color = TYPE_COLORS.get(match["type"])
                if color:
                    cell.fill = _fill(color)
                    cell.font = Font(color="FFFFFFFF")
                if match.get("hours"):
                    cell.value = match["hours"]
                cell.alignment = Alignment(horizontal="center")
                if match["type"] == "ferie":
                    total_ferie_days += 1
            elif hours:
                cell.fill = _fill(PRESENTE_COLOR)
                cell.font = Font(color="FFFFFFFF")
                cell.value = round(hours, 2)
                cell.alignment = Alignment(horizontal="center")
            if hours:
                total_hours += hours
        ws.cell(row=row, column=totals_col, value=round(total_hours, 2))
        ws.cell(row=row, column=ferie_col, value=total_ferie_days)
        row += 1

    legend_row = row + 2
    ws.cell(row=legend_row, column=1, value="Legenda").font = Font(bold=True)
    legend_row += 1
    for key, label in TYPE_LABELS.items():
        swatch = ws.cell(row=legend_row, column=1)
        swatch.fill = _fill(TYPE_COLORS[key])
        ws.cell(row=legend_row, column=2, value=label)
        legend_row += 1
    swatch = ws.cell(row=legend_row, column=1)
    swatch.fill = _fill(PRESENTE_COLOR)
    ws.cell(row=legend_row, column=2, value="Presente")

    _build_dettaglio_sheet(wb, month, leave_requests, sessions)
    return wb


def _build_dettaglio_sheet(wb: Workbook, month: str, leave_requests: list, sessions: list) -> None:
    """Stessa lista piatta (una riga per sessione/assenza) del precedente
    export CSV — vedi attendance_service.export_xlsx, che chiama questa
    funzione tramite build_attendance_workbook."""
    ws = wb.create_sheet("Dettaglio")
    headers = ["Dipendente", "Tipo", "Data", "Data fine", "Entrata", "Uscita", "Ore", "Note"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)

    year, mon = (int(p) for p in month.split("-"))
    month_start = f"{month}-01"
    month_end = f"{month}-{monthrange(year, mon)[1]:02d}"

    rows = []
    for s in sessions:
        day = local_date_str(s["clock_in"])
        hours = (datetime.fromisoformat(s["clock_out"]) - datetime.fromisoformat(s["clock_in"])).total_seconds() / 3600
        rows.append([
            s.get("employee_name", ""), "Presenza", day, "",
            _local_time_str(s["clock_in"]), _local_time_str(s["clock_out"]),
            round(hours, 2), s.get("note", ""),
        ])
    for r in leave_requests:
        rows.append([
            r.get("employee_name", ""), LEAVE_TYPE_LABELS.get(r["type"], r["type"]),
            max(r["date_from"], month_start), min(r["date_to"], month_end),
            "", "", r.get("hours") if r.get("hours") is not None else "", r.get("note", ""),
        ])
    rows.sort(key=lambda r: (r[0], r[2]))

    for row_idx, r in enumerate(rows, start=2):
        for col_idx, value in enumerate(r, start=1):
            ws.cell(row=row_idx, column=col_idx, value=sanitize_cell_text(value))

    widths = [22, 14, 12, 12, 9, 9, 8, 30]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
